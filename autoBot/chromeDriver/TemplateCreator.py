import openpyxl
from random import randint


def create_template(file_name: str, product_code=""):
    if isinstance(file_name, str):
        book = openpyxl.open(file_name, read_only=True)
        sheet = book["ШаблонОтвета"]
        start = sheet[f"A{randint(2, sheet.max_row)}"].value
        gratitude = sheet[f"B{randint(2, sheet.max_row)}"].value
        sheet = book["ОкончаниеОтвета"]
        end = sheet[f"A{randint(2, sheet.max_row)}"].value
        result = start + gratitude + end
        return result
    raise TypeError("Параметр 'file_name' должен быть экземлпяром класса 'str'.")


if __name__ == '__main__':
    print(create_template(file_name="Шаблон.xlsx"))
